#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
EKSTRE MUTABAKAT PROGRAMI
=========================
İki firmanın karşılıklı cari hesap ekstresini (Mikro muhasebe programı
formatında) okuyup satır satır karşılaştırır; eşleşen, yalnız bir tarafta
bulunan ve tutarı uyuşmayan hareketleri tek bir Excel raporunda sunar.

KULLANIM
--------
    python3 ekstre_mutabakat.py <dosya1.xlsx> <dosya2.xls> [-o cikti.xlsx]
                                 [--firma1 "Kendi Firmanız"]
                                 [--firma2 "Karşı Firma"]

Desteklenen iki ekstre formatı otomatik tanınır:
  Format A ("hesap_karti")  -> TARİH, EVRAK TİPİ, ANA DÖVİZ BORÇ, ANA DÖVİZ
                                ALACAK, GIB FATURA NO ... sütunlarını içerir.
  Format B ("cari_ekstre")  -> İşlem Tarihi, İşlem Türü, Borç Tutarı,
                                Alacak Tutarı, Evrak No ... sütunlarını içerir.

Yeni bir formatla karşılaşırsanız aşağıdaki `FORMAT_TANIMLARI` sözlüğüne
birkaç satırlık bir eşleme ekleyerek programı genişletebilirsiniz.

EŞLEŞTİRME MANTIĞI
-------------------
1) Önce evrak no'su aynı olan ve yönü (borç/alacak) ters olan hareketler
   eşleştirilir. Tutar ufak da olsa farklıysa "Tutar Farklı" olarak işaretlenir.
2) Evrak no ile eşleşmeyenler için: yönü ters ve tutarı (kuruşuna kadar)
   aynı olan, tarihi en yakın hareketle greedy (açgözlü) eşleştirme yapılır.
3) Eşi bulunamayan hareketler "Yalnız X'de var" olarak raporlanır.
"""

import argparse
import re
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------
# 1) FORMAT TANIMLARI — yeni ekstre biçimleri buraya eklenir
# --------------------------------------------------------------------------
FORMAT_TANIMLARI = {
    "hesap_karti": {
        "gerekli_sutunlar": {"TARİH", "EVRAK TİPİ", "ANA DÖVİZ BORÇ", "ANA DÖVİZ ALACAK"},
        "tarih_col": "TARİH",
        "tip_col": "EVRAK TİPİ",
        "borc_col": "ANA DÖVİZ BORÇ",
        "alacak_col": "ANA DÖVİZ ALACAK",
        "evrak_col": "GIB FATURA NO",
        "ek_bilgi_col": "SORUMLU İSMİ",
        "tarih_format": None,  # pandas otomatik algılar
    },
    "cari_ekstre": {
        "gerekli_sutunlar": {"İşlem Tarihi", "İşlem Türü", "Borç Tutarı", "Alacak Tutarı"},
        "tarih_col": "İşlem Tarihi",
        "tip_col": "İşlem Türü",
        "borc_col": "Borç Tutarı",
        "alacak_col": "Alacak Tutarı",
        "evrak_col": "Evrak No",
        "ek_bilgi_col": "Açıklama",
        "tarih_format": "%d.%m.%Y %H:%M:%S",
    },
    "logo_ekstre": {
        "gerekli_sutunlar": {"TARIH", "BELGE NO", "BELGETURU", "BORC", "ALACAK"},
        "tarih_col": "TARIH",
        "tip_col": "BELGETURU",
        "borc_col": "BORC",
        "alacak_col": "ALACAK",
        "evrak_col": "BELGE NO",
        "ek_bilgi_col": "ACIKLAMA",
        "tarih_format": None,  # pandas otomatik algılar
    },
    "doviz_ekstre": {
        "gerekli_sutunlar": {"Tarih", "Fiş No", "İşlem Türü", "Borç", "Alacak", "PB", "Kur", "İşlem TL"},
        "tarih_col": "Tarih",
        "tip_col": "İşlem Türü",
        "borc_col": "Borç",
        "alacak_col": "Alacak",
        "evrak_col": "Fiş No",
        "ek_bilgi_col": None,
        "tarih_format": None,  # pandas otomatik algılar
        # Borç/Alacak dövizli sütunlar sadece YÖN belirlemek için kullanılır;
        # gerçek karşılaştırma tutarı her zaman TL karşılığından alınır —
        # aksi halde döviz tutarı TL ekstresiyle asla eşleşmez.
        "tutar_kaynak_col": "İşlem TL",
    },
}


def _parse_sayi(deger):
    """Sayıyı, para birimi sembolü/etiketi ve Türkçe biçim (1.234,56) içerse bile ayrıştırır."""
    if pd.isna(deger):
        return 0.0
    if isinstance(deger, (int, float)):
        return float(deger)
    s = re.sub(r"[^\d,.\-]", "", str(deger))
    if not s:
        return 0.0
    if re.match(r"^-?\d{1,3}(\.\d{3})*(,\d+)?$", s):
        s = s.replace(".", "").replace(",", ".")
    elif "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def dosya_oku(path):
    """xlsx/xls dosyasını pandas ile okur (xls için xlrd otomatik kullanılır)."""
    try:
        return pd.read_excel(path)
    except ImportError as e:
        sys.exit(f"HATA: {e}\nÇözüm: pip install xlrd --break-system-packages")


def format_algila(df, dosya_adi):
    for isim, tanim in FORMAT_TANIMLARI.items():
        if tanim["gerekli_sutunlar"].issubset(set(df.columns)):
            return isim, tanim
    sys.exit(
        f"HATA: '{dosya_adi}' dosyasının formatı tanınamadı.\n"
        f"Sütunlar: {list(df.columns)}\n"
        f"Yeni bir format tanımı eklemeniz gerekiyor (bkz. FORMAT_TANIMLARI)."
    )


def hareketleri_cikar(df, tanim, dosya_kodu):
    """Her satırı, tek yönlü (borç ya da alacak) 'hareket' kayıtlarına çevirir."""
    d = df.copy()

    if tanim["tarih_format"]:
        d["_tarih"] = pd.to_datetime(d[tanim["tarih_col"]], format=tanim["tarih_format"], errors="coerce")
    else:
        d["_tarih"] = pd.to_datetime(d[tanim["tarih_col"]], errors="coerce")

    borc_yon = d[tanim["borc_col"]].map(_parse_sayi)
    alacak_yon = d[tanim["alacak_col"]].map(_parse_sayi)
    tutar_kaynak_col = tanim.get("tutar_kaynak_col")
    if tutar_kaynak_col:
        # Borç/Alacak dövizli sütunlar sadece YÖN belirlemek için kullanılır;
        # gerçek karşılaştırma tutarı her zaman TL karşılığından alınır.
        tutar_tl = d[tutar_kaynak_col].map(_parse_sayi).abs()
        d["_borc"] = tutar_tl.where(borc_yon != 0, 0).round(2)
        d["_alacak"] = tutar_tl.where(alacak_yon != 0, 0).round(2)
    else:
        d["_borc"] = borc_yon.round(2)
        d["_alacak"] = alacak_yon.round(2)
    d["_evrak"] = d[tanim["evrak_col"]] if tanim["evrak_col"] in d.columns else None
    d["_ekbilgi"] = d[tanim["ek_bilgi_col"]] if tanim["ek_bilgi_col"] in d.columns else None
    d["_tip"] = d[tanim["tip_col"]]
    d = d.dropna(subset=["_tarih"])  # devir/toplam gibi tarihsiz satırları at
    d = d[d["_tip"].notna()]
    # Mikro dışa aktarımlarında devir (açılış bakiyesi) satırları "_DEVİR_" gibi
    # işaretlenir; gerçek bir hareket olmadığından mutabakat dışı bırakılır.
    d = d[~d["_tip"].astype(str).str.upper().str.contains("DEVİR|DEVIR")]
    d["_satir_id"] = range(len(d))

    hareketler = []
    for _, r in d.iterrows():
        if r["_borc"] != 0:
            hareketler.append({"satir_id": r["_satir_id"], "tarih": r["_tarih"], "tip": r["_tip"],
                                "yon": "BORC", "tutar": r["_borc"], "evrak": r["_evrak"], "dosya": dosya_kodu})
        if r["_alacak"] != 0:
            hareketler.append({"satir_id": r["_satir_id"], "tarih": r["_tarih"], "tip": r["_tip"],
                                "yon": "ALACAK", "tutar": r["_alacak"], "evrak": r["_evrak"], "dosya": dosya_kodu})

    return d.set_index("_satir_id"), pd.DataFrame(hareketler)


def esletir(h1, h2, evrak_tol=1.0, tutar_tol=0.05):
    """h1 ve h2 hareket tablolarını eşleştirir; h1'e 'matched/match_id2/status' ekler."""
    h1 = h1.copy()
    h2 = h2.copy()
    h1["matched"] = False
    h1["match_id2"] = None
    h1["status"] = None
    h1["grup"] = None
    h2["matched"] = False
    h2["match_id1"] = None
    h2["status"] = None
    h2["grup"] = None

    ters = {"BORC": "ALACAK", "ALACAK": "BORC"}

    # 1. Geçiş: evrak no eşleşmesi
    for i, r1 in h1.iterrows():
        if pd.isna(r1["evrak"]) or r1["evrak"] is None:
            continue
        aday = h2[(~h2["matched"]) & (h2["evrak"] == r1["evrak"]) & (h2["yon"] == ters[r1["yon"]])]
        if aday.empty:
            continue
        aday = aday.copy()
        aday["fark"] = (aday["tutar"] - r1["tutar"]).abs()
        en_iyi = aday.sort_values("fark").iloc[0]
        h1.at[i, "matched"] = True
        h1.at[i, "match_id2"] = en_iyi["satir_id"]
        h1.at[i, "status"] = "Eşleşti" if en_iyi["fark"] <= evrak_tol else "Tutar Farklı"
        h2.loc[h2["satir_id"] == en_iyi["satir_id"], "matched"] = True
        h2.loc[h2["satir_id"] == en_iyi["satir_id"], "match_id1"] = r1["satir_id"]

    # 2. Geçiş: tutar (birebir) + yön + en yakın tarih
    kalan = h1[~h1["matched"]].sort_values("tarih")
    for i, r1 in kalan.iterrows():
        aday = h2[(~h2["matched"]) & (h2["yon"] == ters[r1["yon"]]) & ((h2["tutar"] - r1["tutar"]).abs() <= tutar_tol)]
        if aday.empty:
            continue
        aday = aday.copy()
        aday["gun_farki"] = (aday["tarih"] - r1["tarih"]).abs()
        en_iyi = aday.sort_values("gun_farki").iloc[0]
        h1.at[i, "matched"] = True
        h1.at[i, "match_id2"] = en_iyi["satir_id"]
        h1.at[i, "status"] = "Eşleşti"
        h2.loc[h2["satir_id"] == en_iyi["satir_id"], "matched"] = True
        h2.loc[h2["satir_id"] == en_iyi["satir_id"], "match_id1"] = r1["satir_id"]

    return h1, h2


def _en_iyi_alt_kume(hedef_tutar, havuz, tol, max_boyut):
    """Alt küme toplamı (subset-sum) araması: havuzdaki hangi satırlar toplandığında
    hedef_tutar'a (tol payıyla) ulaşılıyor? Pozitif tutarlar + tutara göre artan sıralama
    ile budama yapan basit bir DFS — küçük mutabakat setleri için yeterince hızlı."""
    havuz = sorted(havuz, key=lambda r: r["tutar"])
    n = len(havuz)
    secili = []

    def dfs(start, toplam):
        if len(secili) >= 2 and abs(toplam - hedef_tutar) <= tol:
            return list(secili)
        if len(secili) >= max_boyut or start >= n:
            return None
        for i in range(start, n):
            r = havuz[i]
            yeni_toplam = toplam + r["tutar"]
            if yeni_toplam > hedef_tutar + tol:
                continue
            secili.append(r)
            sonuc = dfs(i + 1, yeni_toplam)
            if sonuc:
                return sonuc
            secili.pop()
        return None

    return dfs(0, 0)


def grupla_eslestir(h1, h2, tol=0.05, max_boyut=5, tarih_penceresi_gun=60):
    """Bir tarafın tek satırının, karşı tarafta 2+ satıra bölünmüş olabileceği durumları
    (N'e 1 eşleştirme) yakalar. Örn: A'da 3.000 TL tek satır, B'de 1.000+1.000+1.000 üç satır.
    Her iki yönde de (A tek/B çoklu ve B tek/A çoklu) arama yapılır."""
    ters = {"BORC": "ALACAK", "ALACAK": "BORC"}
    grup_no = 0

    def tek_yonden_ara(hedef_df, havuz_df):
        nonlocal grup_no
        hedefler = hedef_df[~hedef_df["matched"]].sort_values("tarih")
        for i, hedef in hedefler.iterrows():
            if hedef_df.at[i, "matched"]:
                continue
            aday_havuz = havuz_df[
                (~havuz_df["matched"])
                & (havuz_df["yon"] == ters[hedef["yon"]])
                & ((havuz_df["tarih"] - hedef["tarih"]).abs() <= pd.Timedelta(days=tarih_penceresi_gun))
            ]
            if len(aday_havuz) < 2:
                continue
            havuz_listesi = aday_havuz.to_dict("records")
            combo = _en_iyi_alt_kume(hedef["tutar"], havuz_listesi, tol, max_boyut)
            if combo:
                grup_no += 1
                hedef_df.at[i, "matched"] = True
                hedef_df.at[i, "status"] = "Gruplu Eşleşme"
                hedef_df.at[i, "grup"] = grup_no
                combo_ids = [c["satir_id"] for c in combo]
                havuz_df.loc[havuz_df["satir_id"].isin(combo_ids), "matched"] = True
                havuz_df.loc[havuz_df["satir_id"].isin(combo_ids), "grup"] = grup_no

    tek_yonden_ara(h1, h2)
    tek_yonden_ara(h2, h1)
    return h1, h2


def rapor_olustur(d1, h1, d2, h2, firma1, firma2, cikti_yolu):
    rows = []
    for _, r1 in h1.iterrows():
        if r1["matched"] and r1["match_id2"] is not None:
            a = d1.loc[r1["satir_id"]]
            b = d2.loc[r1["match_id2"]]
            rows.append({
                "T1": a["_tarih"], "Tip1": a["_tip"], "Borc1": a["_borc"], "Alacak1": a["_alacak"],
                "Evrak1": a["_evrak"], "Ek1": a["_ekbilgi"],
                "T2": b["_tarih"], "Tip2": b["_tip"], "Borc2": b["_borc"], "Alacak2": b["_alacak"],
                "Evrak2": b["_evrak"], "Ek2": b["_ekbilgi"],
                "Durum": r1["status"],
            })
        elif r1["status"] == "Gruplu Eşleşme":
            a = d1.loc[r1["satir_id"]]
            rows.append({
                "T1": a["_tarih"], "Tip1": a["_tip"], "Borc1": a["_borc"], "Alacak1": a["_alacak"],
                "Evrak1": a["_evrak"], "Ek1": a["_ekbilgi"],
                "T2": None, "Tip2": None, "Borc2": None, "Alacak2": None, "Evrak2": None, "Ek2": None,
                "Durum": f"Gruplu Eşleşme · Grup {int(r1['grup'])}",
            })
        elif not r1["matched"]:
            a = d1.loc[r1["satir_id"]]
            rows.append({
                "T1": a["_tarih"], "Tip1": a["_tip"], "Borc1": a["_borc"], "Alacak1": a["_alacak"],
                "Evrak1": a["_evrak"], "Ek1": a["_ekbilgi"],
                "T2": None, "Tip2": None, "Borc2": None, "Alacak2": None, "Evrak2": None, "Ek2": None,
                "Durum": f"Yalnız {firma1}'de var",
            })
    for _, r2 in h2.iterrows():
        if r2["matched"] and r2["match_id1"] is not None:
            continue  # h1 döngüsünde eşli satır olarak zaten eklendi
        if r2["matched"] and r2.get("grup") is not None:
            b = d2.loc[r2["satir_id"]]
            rows.append({
                "T1": None, "Tip1": None, "Borc1": None, "Alacak1": None, "Evrak1": None, "Ek1": None,
                "T2": b["_tarih"], "Tip2": b["_tip"], "Borc2": b["_borc"], "Alacak2": b["_alacak"],
                "Evrak2": b["_evrak"], "Ek2": b["_ekbilgi"],
                "Durum": f"Gruplu Eşleşme · Grup {int(r2['grup'])}",
            })
        elif not r2["matched"]:
            b = d2.loc[r2["satir_id"]]
            rows.append({
                "T1": None, "Tip1": None, "Borc1": None, "Alacak1": None, "Evrak1": None, "Ek1": None,
                "T2": b["_tarih"], "Tip2": b["_tip"], "Borc2": b["_borc"], "Alacak2": b["_alacak"],
                "Evrak2": b["_evrak"], "Ek2": b["_ekbilgi"],
                "Durum": f"Yalnız {firma2}'de var",
            })

    merged = pd.DataFrame(rows)
    merged["_sirala"] = merged["T1"].combine_first(merged["T2"])
    merged = merged.sort_values("_sirala").reset_index(drop=True)

    # ---- Excel yazımı (tek sayfa) ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Karşılaştırmalı Ekstre"

    title_font = Font(name="Arial", bold=True, size=13)
    sub_font = Font(name="Arial", italic=True, size=10, color="555555")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    normal_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", size=10, bold=True)

    fill_f1 = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    fill_f2 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    fill_durum_h = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")

    fill_match = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_only1 = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fill_only2 = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_diff = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_grup = PatternFill(start_color="EAE6F5", end_color="EAE6F5", fill_type="solid")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sayilar = merged["Durum"].value_counts()
    ozet = "  |  ".join(f"{k}: {v}" for k, v in sayilar.items())

    ws.merge_cells("A1:M1")
    ws["A1"] = f"CARİ HESAP EKSTRESİ — {firma1.upper()} ile {firma2.upper()} KARŞILAŞTIRMASI"
    ws["A1"].font = title_font
    ws.merge_cells("A2:M2")
    ws["A2"] = f"Toplam {len(merged)} satır  |  {ozet}     (Rapor tarihi: {datetime.now().strftime('%d.%m.%Y')})"
    ws["A2"].font = sub_font

    ws.merge_cells("A4:F4")
    ws["A4"] = firma1.upper()
    ws["A4"].font = header_font
    ws["A4"].fill = fill_f1
    ws["A4"].alignment = Alignment(horizontal="center")

    ws.merge_cells("G4:L4")
    ws["G4"] = firma2.upper()
    ws["G4"].font = header_font
    ws["G4"].fill = fill_f2
    ws["G4"].alignment = Alignment(horizontal="center")

    ws.cell(row=4, column=13, value="DURUM").font = header_font
    ws.cell(row=4, column=13).fill = fill_durum_h
    ws.cell(row=4, column=13).alignment = Alignment(horizontal="center")

    basliklar = [
        ("Tarih", fill_f1), ("Evrak Tipi", fill_f1), ("Borç (TL)", fill_f1), ("Alacak (TL)", fill_f1),
        ("Evrak No", fill_f1), ("Ek Bilgi", fill_f1),
        ("Tarih", fill_f2), ("Evrak Tipi", fill_f2), ("Borç (TL)", fill_f2), ("Alacak (TL)", fill_f2),
        ("Evrak No", fill_f2), ("Ek Bilgi", fill_f2),
        ("Durum", fill_durum_h),
    ]
    for j, (ad, fill) in enumerate(basliklar, start=1):
        c = ws.cell(row=5, column=j, value=ad)
        c.font = header_font
        c.fill = fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = border

    kolonlar = ["T1", "Tip1", "Borc1", "Alacak1", "Evrak1", "Ek1",
                "T2", "Tip2", "Borc2", "Alacak2", "Evrak2", "Ek2", "Durum"]

    durum_fill = {
        "Eşleşti": fill_match,
        f"Yalnız {firma1}'de var": fill_only1,
        f"Yalnız {firma2}'de var": fill_only2,
        "Tutar Farklı": fill_diff,
    }

    for i, (_, row) in enumerate(merged.iterrows(), start=6):
        durum_str = str(row["Durum"])
        if durum_str.startswith("Gruplu Eşleşme"):
            fill = fill_grup
        else:
            fill = durum_fill.get(row["Durum"])
        for j, col in enumerate(kolonlar, start=1):
            val = row[col]
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            c = ws.cell(row=i, column=j, value=val if pd.notna(val) else None)
            c.font = bold_font if col == "Durum" else normal_font
            c.border = border
            if fill:
                c.fill = fill
            if col in ("T1", "T2"):
                c.number_format = "DD.MM.YYYY"
            if col in ("Borc1", "Alacak1", "Borc2", "Alacak2"):
                c.number_format = "#,##0.00"

    genislikler = [12, 20, 12, 12, 17, 22, 12, 20, 12, 12, 17, 22, 20]
    for j, w in enumerate(genislikler, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws.freeze_panes = ws.cell(row=6, column=1)
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28

    wb.save(cikti_yolu)
    return merged, cikti_yolu


def main():
    ap = argparse.ArgumentParser(description="İki ekstre dosyasını karşılaştırıp tek sayfalık mutabakat raporu üretir.")
    ap.add_argument("dosya1", help="Kendi firmanızın ekstre dosyası (.xlsx/.xls)")
    ap.add_argument("dosya2", help="Karşı firmanın ekstre dosyası (.xlsx/.xls)")
    ap.add_argument("-o", "--cikti", default="mutabakat_raporu.xlsx", help="Çıktı Excel dosyasının yolu")
    ap.add_argument("--firma1", required=True, help="Kendi firmanızın adı (rapor başlığında ve sütunlarında kullanılır)")
    ap.add_argument("--firma2", required=True, help="Karşı firmanın adı (rapor başlığında ve sütunlarında kullanılır)")
    args = ap.parse_args()

    df1_raw = dosya_oku(args.dosya1)
    df2_raw = dosya_oku(args.dosya2)

    _, tanim1 = format_algila(df1_raw, args.dosya1)
    _, tanim2 = format_algila(df2_raw, args.dosya2)

    d1, h1 = hareketleri_cikar(df1_raw, tanim1, "F1")
    d2, h2 = hareketleri_cikar(df2_raw, tanim2, "F2")

    h1, h2 = esletir(h1, h2)
    h1, h2 = grupla_eslestir(h1, h2)

    merged, out_path = rapor_olustur(d1, h1, d2, h2, args.firma1, args.firma2, args.cikti)

    print(f"Toplam {len(merged)} satır işlendi.")
    print(merged["Durum"].value_counts().to_string())
    print(f"\nRapor kaydedildi: {out_path}")


if __name__ == "__main__":
    main()
